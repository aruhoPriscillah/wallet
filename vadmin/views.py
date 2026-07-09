import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.contrib import messages
from datetime import timedelta
from wallet.models import (Wallet, Transaction, KYCDocument,
                           WithdrawalRequest, FraudFlag, WalletFreeze)


def admin_required(view_func):
    return staff_member_required(view_func, login_url='/login/')


@admin_required
def admin_dashboard(request):
    now = timezone.now()
    today = now.date()
    week_ago = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)

    # Stats
    total_users = User.objects.count()
    new_users_today = User.objects.filter(date_joined__date=today).count()
    new_users_week = User.objects.filter(date_joined__gte=week_ago).count()

    total_wallets = Wallet.objects.count()
    total_balance = Wallet.objects.aggregate(t=Sum('balance'))['t'] or 0
    frozen_wallets = WalletFreeze.objects.filter(is_active=True).count()

    total_transactions = Transaction.objects.count()
    transactions_today = Transaction.objects.filter(created_at__date=today).count()
    volume_today = Transaction.objects.filter(
        created_at__date=today, status='completed'
    ).aggregate(t=Sum('amount'))['t'] or 0
    volume_month = Transaction.objects.filter(
        created_at__gte=month_ago, status='completed'
    ).aggregate(t=Sum('amount'))['t'] or 0

    pending_kyc = KYCDocument.objects.filter(status='pending').count()
    pending_withdrawals = WithdrawalRequest.objects.filter(status='pending').count()
    active_flags = FraudFlag.objects.filter(resolved=False).count()
    high_flags = FraudFlag.objects.filter(resolved=False, severity='high').count()

    recent_transactions = Transaction.objects.select_related(
        'wallet__user'
    ).all()[:10]

    recent_flags = FraudFlag.objects.filter(
        resolved=False
    ).select_related('wallet__user').order_by('-created_at')[:5]

    # Daily transaction volume for chart (last 7 days)
    chart_data = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        vol = Transaction.objects.filter(
            created_at__date=day, status='completed'
        ).aggregate(t=Sum('amount'))['t'] or 0
        chart_data.append({'date': day.strftime('%b %d'), 'amount': float(vol)})

    return render(request, 'vadmin/dashboard.html', {
        'total_users': total_users,
        'new_users_today': new_users_today,
        'new_users_week': new_users_week,
        'total_balance': total_balance,
        'frozen_wallets': frozen_wallets,
        'total_transactions': total_transactions,
        'transactions_today': transactions_today,
        'volume_today': volume_today,
        'volume_month': volume_month,
        'pending_kyc': pending_kyc,
        'pending_withdrawals': pending_withdrawals,
        'active_flags': active_flags,
        'high_flags': high_flags,
        'recent_transactions': recent_transactions,
        'recent_flags': recent_flags,
        'chart_data': chart_data,
    })



@admin_required
def manage_users(request):
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    users = User.objects.select_related('wallet').order_by('-date_joined')
    if query:
        users = users.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )
    if status == 'frozen':
        users = users.filter(wallet__freeze__is_active=True)
    elif status == 'active':
        users = users.exclude(wallet__freeze__is_active=True)

    return render(request, 'vadmin/users.html', {
        'users': users, 'query': query, 'status': status
    })


@admin_required
def user_detail(request, user_id):
    user = get_object_or_404(User, id=user_id)
    wallet, _ = Wallet.objects.get_or_create(user=user)
    transactions = wallet.transactions.all()[:20]
    try:
        kyc = wallet.kyc
    except KYCDocument.DoesNotExist:
        kyc = None
    try:
        freeze = wallet.freeze
        is_frozen = freeze.is_active
    except WalletFreeze.DoesNotExist:
        freeze = None
        is_frozen = False
    flags = wallet.fraud_flags.all()
    return render(request, 'vadmin/user_detail.html', {
        'viewed_user': user, 'wallet': wallet, 'transactions': transactions,
        'kyc': kyc, 'freeze': freeze, 'is_frozen': is_frozen, 'flags': flags,
    })


@admin_required
def freeze_wallet(request, wallet_id):
    wallet = get_object_or_404(Wallet, id=wallet_id)
    if request.method == 'POST':
        reason = request.POST.get('reason', 'Suspicious activity')
        freeze, created = WalletFreeze.objects.get_or_create(wallet=wallet)
        freeze.reason = reason
        freeze.frozen_by = request.user
        freeze.is_active = True
        freeze.save()
        messages.success(request, f'Wallet for @{wallet.user.username} has been frozen.')
        return redirect('admin_user_detail', user_id=wallet.user.id)
    return redirect('admin_user_detail', user_id=wallet.user.id)


@admin_required
def unfreeze_wallet(request, wallet_id):
    wallet = get_object_or_404(Wallet, id=wallet_id)
    try:
        wallet.freeze.is_active = False
        wallet.freeze.save()
        messages.success(request, f'Wallet for @{wallet.user.username} has been unfrozen.')
    except WalletFreeze.DoesNotExist:
        pass
    return redirect('admin_user_detail', user_id=wallet.user.id)


@admin_required
def kyc_list(request):
    status = request.GET.get('status', 'pending')
    docs = KYCDocument.objects.select_related(
        'wallet__user'
    ).filter(status=status).order_by('-submitted_at')
    return render(request, 'vadmin/kyc.html', {'docs': docs, 'status': status})


@admin_required
def kyc_review(request, doc_id):
    doc = get_object_or_404(KYCDocument, id=doc_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        doc.notes = notes
        doc.reviewed_at = timezone.now()
        if action == 'approve':
            doc.status = 'approved'
            messages.success(request, f'KYC approved for @{doc.wallet.user.username}.')
        elif action == 'reject':
            doc.status = 'rejected'
            messages.warning(request, f'KYC rejected for @{doc.wallet.user.username}.')
        doc.save()
        return redirect('admin_kyc')
    return render(request, 'vadmin/kyc_review.html', {'doc': doc})


@admin_required
def withdrawals(request):
    status = request.GET.get('status', 'pending')
    reqs = WithdrawalRequest.objects.select_related(
        'wallet__user'
    ).filter(status=status).order_by('-requested_at')
    return render(request, 'vadmin/withdrawals.html', {'requests': reqs, 'status': status})


@admin_required
def withdrawal_review(request, req_id):
    req = get_object_or_404(WithdrawalRequest, id=req_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('admin_notes', '')
        req.admin_notes = notes
        req.reviewed_at = timezone.now()
        if action == 'approve':
            req.status = 'approved'
            messages.success(request, f'Withdrawal of UGX {req.amount:,.0f} approved.')
        elif action == 'reject':
            req.status = 'rejected'
            req.wallet.balance += req.amount
            req.wallet.save()
            messages.warning(request, f'Withdrawal rejected. Balance refunded.')
        req.save()
        return redirect('admin_withdrawals')
    return render(request, 'vadmin/withdrawal_review.html', {'req': req})


@admin_required
def transactions(request):
    txn_type = request.GET.get('type', '')
    query = request.GET.get('q', '')
    txns = Transaction.objects.select_related('wallet__user').order_by('-created_at')
    if txn_type:
        txns = txns.filter(transaction_type=txn_type)
    if query:
        txns = txns.filter(
            Q(wallet__user__username__icontains=query) |
            Q(description__icontains=query)
        )
    return render(request, 'vadmin/transactions.html', {
        'transactions': txns[:100],
        'filter_type': txn_type,
        'query': query,
    })

@admin_required
def fraud_monitoring(request):
    severity = request.GET.get('severity', '')
    flags = FraudFlag.objects.select_related('wallet__user').filter(resolved=False)
    if severity:
        flags = flags.filter(severity=severity)
    flags = flags.order_by('-created_at')
    return render(request, 'vadmin/fraud.html', {'flags': flags, 'severity': severity})


@admin_required
def resolve_flag(request, flag_id):
    flag = get_object_or_404(FraudFlag, id=flag_id)
    flag.resolved = True
    flag.save()
    messages.success(request, 'Fraud flag resolved.')
    return redirect('admin_fraud')


@admin_required
def add_fraud_flag(request, wallet_id):
    wallet = get_object_or_404(Wallet, id=wallet_id)
    if request.method == 'POST':
        reason = request.POST.get('reason')
        severity = request.POST.get('severity', 'low')
        FraudFlag.objects.create(wallet=wallet, reason=reason, severity=severity)
        messages.success(request, f'Fraud flag added for @{wallet.user.username}.')
    return redirect('admin_user_detail', user_id=wallet.user.id)


@admin_required
def reports(request):
    from django.db.models.functions import TruncDay, TruncMonth
    now = timezone.now()
    today = now.date()
    month_ago = now - timedelta(days=30)
    year_ago = now - timedelta(days=365)

    # Summary totals (30 days)
    deposit_total = Transaction.objects.filter(
        transaction_type='deposit', status='completed', created_at__gte=month_ago
    ).aggregate(t=Sum('amount'))['t'] or 0

    withdrawal_total = Transaction.objects.filter(
        transaction_type='withdrawal', status='completed', created_at__gte=month_ago
    ).aggregate(t=Sum('amount'))['t'] or 0

    transfer_total = Transaction.objects.filter(
        transaction_type='transfer_out', status='completed', created_at__gte=month_ago
    ).aggregate(t=Sum('amount'))['t'] or 0

    total_revenue = deposit_total  # adjust if you charge fees

    # Active users (made at least 1 transaction in last 30 days)
    active_users = Wallet.objects.filter(
        transactions__created_at__gte=month_ago,
        transactions__status='completed'
    ).distinct().count()

    # Daily transaction volume (last 30 days)
    daily_volume = Transaction.objects.filter(
        status='completed', created_at__gte=month_ago
    ).annotate(day=TruncDay('created_at')).values('day').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('day')

    daily_chart = [
        {
            'date': entry['day'].strftime('%b %d'),
            'amount': float(entry['total']),
            'count': entry['count']
        }
        for entry in daily_volume
    ]

    # Monthly income (last 12 months)
    monthly_volume = Transaction.objects.filter(
        status='completed', created_at__gte=year_ago
    ).annotate(month=TruncMonth('created_at')).values('month').annotate(
        deposits=Sum('amount', filter=Q(transaction_type='deposit')),
        withdrawals=Sum('amount', filter=Q(transaction_type='withdrawal')),
        transfers=Sum('amount', filter=Q(transaction_type='transfer_out')),
        count=Count('id')
    ).order_by('month')

    monthly_chart = [
        {
            'month': entry['month'].strftime('%b %Y'),
            'deposits': float(entry['deposits'] or 0),
            'withdrawals': float(entry['withdrawals'] or 0),
            'transfers': float(entry['transfers'] or 0),
            'count': entry['count'],
        }
        for entry in monthly_volume
    ]

    # Top users by transaction volume
    top_users = Wallet.objects.annotate(
        txn_count=Count('transactions', filter=Q(transactions__status='completed')),
        txn_volume=Sum('transactions__amount', filter=Q(transactions__status='completed'))
    ).filter(txn_volume__isnull=False).order_by('-txn_volume')[:10]

    # Transaction type breakdown
    type_breakdown = Transaction.objects.filter(
        status='completed', created_at__gte=month_ago
    ).values('transaction_type').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('-total')

    type_chart = [
        {'type': entry['transaction_type'], 'total': float(entry['total']), 'count': entry['count']}
        for entry in type_breakdown
    ]

    # Daily active users (last 30 days)
    daily_users = Transaction.objects.filter(
        status='completed', created_at__gte=month_ago
    ).annotate(day=TruncDay('created_at')).values('day').annotate(
        users=Count('wallet', distinct=True)
    ).order_by('day')

    daily_users_chart = [
        {'date': entry['day'].strftime('%b %d'), 'users': entry['users']}
        for entry in daily_users
    ]

    return render(request, 'vadmin/reports.html', {
        'deposit_total': deposit_total,
        'withdrawal_total': withdrawal_total,
        'transfer_total': transfer_total,
        'total_revenue': total_revenue,
        'active_users': active_users,
        'top_users': top_users,
        'daily_chart': json.dumps(daily_chart),
        'monthly_chart': json.dumps(monthly_chart),
        'type_chart': json.dumps(type_chart),
        'daily_users_chart': json.dumps(daily_users_chart),
        'type_breakdown': type_breakdown,
    })

@admin_required
def export_report(request):
    report_type = request.GET.get('type', 'transactions')
    month_ago = timezone.now() - timedelta(days=30)

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='7C6AF7', end_color='7C6AF7', fill_type='solid')
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal='center')

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    if report_type == 'transactions':
        ws = wb.active
        ws.title = 'Transactions'

        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = 'Vault — Transaction Report (Last 30 Days)'
        ws['A1'].font = title_font
        ws['A1'].alignment = center

        ws.merge_cells('A2:G2')
        ws['A2'] = f'Generated: {timezone.now().strftime("%B %d, %Y %H:%M")}'
        ws['A2'].alignment = center

        # Headers
        headers = ['#', 'Username', 'Type', 'Amount (UGX)', 'Description', 'Status', 'Date']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        style_header_row(ws, 4, len(headers))

        # Data
        txns = Transaction.objects.select_related('wallet__user').filter(
            created_at__gte=month_ago
        ).order_by('-created_at')

        for row, txn in enumerate(txns, 5):
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=txn.wallet.user.username)
            ws.cell(row=row, column=3, value=txn.transaction_type)
            ws.cell(row=row, column=4, value=float(txn.amount))
            ws.cell(row=row, column=5, value=txn.description)
            ws.cell(row=row, column=6, value=txn.status)
            ws.cell(row=row, column=7, value=txn.created_at.strftime('%Y-%m-%d %H:%M'))

        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20

    elif report_type == 'users':
        ws = wb.active
        ws.title = 'Users'

        ws.merge_cells('A1:F1')
        ws['A1'] = 'Vault — Users Report'
        ws['A1'].font = title_font
        ws['A1'].alignment = center

        ws.merge_cells('A2:F2')
        ws['A2'] = f'Generated: {timezone.now().strftime("%B %d, %Y %H:%M")}'
        ws['A2'].alignment = center

        headers = ['#', 'Username', 'Full Name', 'Email', 'Balance (UGX)', 'Joined']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        style_header_row(ws, 4, len(headers))

        users = User.objects.select_related('wallet').order_by('-date_joined')
        for row, user in enumerate(users, 5):
            ws.cell(row=row, column=1, value=row - 4)
            ws.cell(row=row, column=2, value=user.username)
            ws.cell(row=row, column=3, value=user.get_full_name())
            ws.cell(row=row, column=4, value=user.email)
            try:
                ws.cell(row=row, column=5, value=float(user.wallet.balance))
            except:
                ws.cell(row=row, column=5, value=0)
            ws.cell(row=row, column=6, value=user.date_joined.strftime('%Y-%m-%d'))

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15

    elif report_type == 'summary':
        ws = wb.active
        ws.title = 'Summary'

        ws.merge_cells('A1:C1')
        ws['A1'] = 'Vault — Monthly Summary Report'
        ws['A1'].font = title_font
        ws['A1'].alignment = center

        ws.merge_cells('A2:C2')
        ws['A2'] = f'Generated: {timezone.now().strftime("%B %d, %Y %H:%M")}'
        ws['A2'].alignment = center

        # Summary stats
        ws.cell(row=4, column=1, value='Metric')
        ws.cell(row=4, column=2, value='Value (UGX)')
        ws.cell(row=4, column=3, value='Period')
        style_header_row(ws, 4, 3)

        deposit_total = Transaction.objects.filter(
            transaction_type='deposit', status='completed', created_at__gte=month_ago
        ).aggregate(t=Sum('amount'))['t'] or 0

        withdrawal_total = Transaction.objects.filter(
            transaction_type='withdrawal', status='completed', created_at__gte=month_ago
        ).aggregate(t=Sum('amount'))['t'] or 0

        transfer_total = Transaction.objects.filter(
            transaction_type='transfer_out', status='completed', created_at__gte=month_ago
        ).aggregate(t=Sum('amount'))['t'] or 0

        total_txns = Transaction.objects.filter(created_at__gte=month_ago).count()
        total_users = User.objects.count()
        active_users = Wallet.objects.filter(
            transactions__created_at__gte=month_ago
        ).distinct().count()
        total_balance = Wallet.objects.aggregate(t=Sum('balance'))['t'] or 0

        rows = [
            ('Total Deposits', float(deposit_total), 'Last 30 days'),
            ('Total Withdrawals', float(withdrawal_total), 'Last 30 days'),
            ('Total Transfers', float(transfer_total), 'Last 30 days'),
            ('Total Transactions', total_txns, 'Last 30 days'),
            ('Total Users', total_users, 'All time'),
            ('Active Users', active_users, 'Last 30 days'),
            ('Total Wallet Balance', float(total_balance), 'Current'),
        ]

        for row_num, (metric, value, period) in enumerate(rows, 5):
            ws.cell(row=row_num, column=1, value=metric)
            ws.cell(row=row_num, column=2, value=value)
            ws.cell(row=row_num, column=3, value=period)

        # Top users sheet
        ws2 = wb.create_sheet('Top Users')
        ws2.merge_cells('A1:D1')
        ws2['A1'] = 'Top Users by Transaction Volume'
        ws2['A1'].font = title_font
        ws2['A1'].alignment = center

        headers2 = ['#', 'Username', 'Transaction Count', 'Total Volume (UGX)']
        for col, header in enumerate(headers2, 1):
            ws2.cell(row=3, column=col, value=header)
        style_header_row(ws2, 3, len(headers2))

        top_users = Wallet.objects.annotate(
            txn_count=Count('transactions'),
            txn_volume=Sum('transactions__amount')
        ).filter(txn_volume__isnull=False).order_by('-txn_volume')[:20]

        for row, w in enumerate(top_users, 4):
            ws2.cell(row=row, column=1, value=row - 3)
            ws2.cell(row=row, column=2, value=w.user.username)
            ws2.cell(row=row, column=3, value=w.txn_count or 0)
            ws2.cell(row=row, column=4, value=float(w.txn_volume or 0))

        ws2.column_dimensions['A'].width = 6
        ws2.column_dimensions['B'].width = 22
        ws2.column_dimensions['C'].width = 22
        ws2.column_dimensions['D'].width = 22

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 18

    # File response
    filename = f'vault_{report_type}_report_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response